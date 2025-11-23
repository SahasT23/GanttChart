# from fastapi import FastAPI, HTTPException, Request, UploadFile, File
# from fastapi.staticfiles import StaticFiles
# from fastapi.responses import FileResponse, Response
# from pydantic import BaseModel
# from typing import List, Optional, Dict
# from datetime import datetime, timedelta
# import uuid
# import io

# import database as db

# app = FastAPI(title="Gantt Chart API", description="Multi-project Gantt chart application with notes")

# # Data models
# class Project(BaseModel):
#     id: str
#     name: str
#     description: Optional[str] = None
#     start_date: Optional[str] = None
#     end_date: Optional[str] = None
#     color: str = "#4285f4"
#     is_active: bool = True
#     created_at: str
#     updated_at: str

# class ProjectCreate(BaseModel):
#     name: str
#     description: Optional[str] = None
#     start_date: Optional[str] = None
#     end_date: Optional[str] = None
#     color: str = "#4285f4"

# class ProjectUpdate(BaseModel):
#     name: Optional[str] = None
#     description: Optional[str] = None
#     start_date: Optional[str] = None
#     end_date: Optional[str] = None
#     color: Optional[str] = None

# class Note(BaseModel):
#     id: str
#     project_id: str
#     note_date: str
#     content: str
#     created_at: str
#     updated_at: str

# class NoteCreate(BaseModel):
#     note_date: str
#     content: str

# class NoteUpdate(BaseModel):
#     content: str

# class Task(BaseModel):
#     id: str
#     project_id: str
#     name: str
#     start_date: str
#     end_date: str
#     progress: float = 0.0
#     color: str = "#4285f4"
#     dependencies: List[str] = []
#     is_milestone: bool = False
#     parent_id: Optional[str] = None
#     created_at: str
#     updated_at: str
#     description: Optional[str] = None
#     assigned_to: Optional[str] = None
#     priority: str = "medium"

# class TaskCreate(BaseModel):
#     name: str
#     start_date: str
#     end_date: str
#     progress: float = 0.0
#     color: str = "#4285f4"
#     dependencies: List[str] = []
#     is_milestone: bool = False
#     parent_id: Optional[str] = None
#     description: Optional[str] = None
#     assigned_to: Optional[str] = None
#     priority: str = "medium"

# class TaskUpdate(BaseModel):
#     name: Optional[str] = None
#     start_date: Optional[str] = None
#     end_date: Optional[str] = None
#     progress: Optional[float] = None
#     color: Optional[str] = None
#     dependencies: Optional[List[str]] = None
#     is_milestone: Optional[bool] = None
#     description: Optional[str] = None
#     assigned_to: Optional[str] = None
#     priority: Optional[str] = None
#     parent_id: Optional[str] = None

# class WeeklyPlanner(BaseModel):
#     id: str
#     project_id: Optional[str] = None
#     week_start_date: str
#     week_end_date: str
#     custom_rows: List[str] = []
#     custom_columns: List[str] = []
#     created_at: str
#     updated_at: str

# class PlannerCreate(BaseModel):
#     week_start_date: str
#     project_id: Optional[str] = None

# class PlannerUpdate(BaseModel):
#     custom_rows: Optional[List[str]] = None
#     custom_columns: Optional[List[str]] = None

# class TimeBlock(BaseModel):
#     id: str
#     planner_id: str
#     day_index: int
#     time_slot: str
#     title: Optional[str] = None
#     description: Optional[str] = None
#     color: str = "#4285f4"
#     created_at: str
#     updated_at: str

# class TimeBlockCreate(BaseModel):
#     day_index: int
#     time_slot: str
#     title: Optional[str] = None
#     description: Optional[str] = None
#     color: str = "#4285f4"

# class TimeBlockUpdate(BaseModel):
#     title: Optional[str] = None
#     description: Optional[str] = None
#     color: Optional[str] = None

# @app.on_event("startup")
# async def startup_event():
#     """Initialize the database when the app starts"""
#     db.init_database()
#     print("✓ Database ready")

# def get_current_project_id():
#     """Get the currently active project ID"""
#     project = db.get_active_project()
#     if project:
#         return project['id']
    
#     # If no active project, get the first one or create default
#     projects = db.get_all_projects()
#     if projects:
#         db.set_active_project(projects[0]['id'])
#         return projects[0]['id']
    
#     # Create default project
#     now = datetime.now().isoformat()
#     default_project = {
#         'id': str(uuid.uuid4()),
#         'name': 'My First Project',
#         'description': 'Default project',
#         'created_at': now,
#         'updated_at': now
#     }
#     db.create_project(default_project)
#     db.set_active_project(default_project['id'])
#     return default_project['id']

# def validate_dependencies(task_id: str, dependencies: List[str], project_id: str) -> List[str]:
#     """Validate and filter dependencies to prevent circular references"""
#     valid_deps = []
#     all_tasks = {task['id']: task for task in db.get_all_tasks(project_id)}
    
#     def has_circular_dependency(current_id: str, target_id: str, visited: set = None) -> bool:
#         if visited is None:
#             visited = set()
        
#         if current_id in visited:
#             return True
        
#         if current_id == target_id:
#             return True
        
#         visited.add(current_id)
        
#         current_task = all_tasks.get(current_id)
#         if current_task:
#             for dep in current_task['dependencies']:
#                 if has_circular_dependency(dep, target_id, visited.copy()):
#                     return True
        
#         return False
    
#     for dep_id in dependencies:
#         if dep_id in all_tasks and not has_circular_dependency(dep_id, task_id):
#             valid_deps.append(dep_id)
    
#     return valid_deps

# def calculate_critical_path(project_id: str) -> List[str]:
#     """Calculate the critical path through the project"""
#     all_tasks = {task['id']: task for task in db.get_all_tasks(project_id)}
    
#     def get_chain_length(task_id: str, visited: set = None) -> int:
#         if visited is None:
#             visited = set()
        
#         if task_id in visited:
#             return 0
        
#         visited.add(task_id)
#         task = all_tasks.get(task_id)
#         if not task:
#             return 0
        
#         if not task['dependencies']:
#             return 1
        
#         max_length = 0
#         for dep_id in task['dependencies']:
#             length = get_chain_length(dep_id, visited.copy())
#             max_length = max(max_length, length)
        
#         return max_length + 1
    
#     max_length = 0
#     critical_task = None
    
#     for task_id in all_tasks:
#         length = get_chain_length(task_id)
#         if length > max_length:
#             max_length = length
#             critical_task = task_id
    
#     critical_path = []
#     if critical_task:
#         def trace_path(task_id: str):
#             critical_path.append(task_id)
#             task = all_tasks.get(task_id)
#             if task and task['dependencies']:
#                 longest_dep = None
#                 max_dep_length = 0
#                 for dep_id in task['dependencies']:
#                     dep_length = get_chain_length(dep_id)
#                     if dep_length > max_dep_length:
#                         max_dep_length = dep_length
#                         longest_dep = dep_id
                
#                 if longest_dep:
#                     trace_path(longest_dep)
        
#         trace_path(critical_task)
    
#     return critical_path

# def log_action(action: str, task_id: str, task_name: str, details: Dict, project_id: str, user: str = "system"):
#     """Add an action to the log"""
#     log_data = {
#         'id': str(uuid.uuid4()),
#         'project_id': project_id,
#         'action': action,
#         'task_id': task_id,
#         'task_name': task_name,
#         'timestamp': datetime.now().isoformat(),
#         'details': details,
#         'user': user
#     }
#     db.create_log(log_data)

# @app.middleware("http")
# async def add_cors_header(request: Request, call_next):
#     response = await call_next(request)
#     response.headers["Access-Control-Allow-Origin"] = "*"
#     response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
#     response.headers["Access-Control-Allow-Headers"] = "*"
#     return response

# # Project endpoints
# @app.get("/api/projects")
# async def get_projects():
#     """Get all projects"""
#     projects = db.get_all_projects()
#     return {"projects": projects}

# @app.get("/api/projects/active")
# async def get_active_project():
#     """Get the currently active project"""
#     project = db.get_active_project()
#     if not project:
#         raise HTTPException(status_code=404, detail="No active project")
#     return {"project": project}

# @app.post("/api/projects")
# async def create_project(project_data: ProjectCreate):
#     """Create a new project"""
#     project_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     project_dict = {
#         'id': project_id,
#         'name': project_data.name,
#         'description': project_data.description,
#         'start_date': project_data.start_date,
#         'end_date': project_data.end_date,
#         'color': project_data.color,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_project(project_dict)
#     db.set_active_project(project_id)
    
#     return {"project": project_dict, "message": "Project created successfully"}

# @app.put("/api/projects/{project_id}")
# async def update_project(project_id: str, updates: ProjectUpdate):
#     """Update an existing project"""
#     project = db.get_project_by_id(project_id)
#     if not project:
#         raise HTTPException(status_code=404, detail="Project not found")
    
#     update_data = updates.model_dump(exclude_unset=True)
#     updated_project = db.update_project(project_id, update_data)
    
#     return {"project": updated_project, "message": "Project updated successfully"}

# @app.delete("/api/projects/{project_id}")
# async def delete_project(project_id: str):
#     """Delete a project and all its data"""
#     project = db.get_project_by_id(project_id)
#     if not project:
#         raise HTTPException(status_code=404, detail="Project not found")
    
#     # Check if it's the only project
#     all_projects = db.get_all_projects()
#     if len(all_projects) == 1:
#         raise HTTPException(status_code=400, detail="Cannot delete the only project")
    
#     db.delete_project(project_id)
    
#     # Set another project as active if this was the active one
#     if project.get('is_active'):
#         remaining_projects = db.get_all_projects()
#         if remaining_projects:
#             db.set_active_project(remaining_projects[0]['id'])
    
#     return {"message": "Project deleted successfully"}

# @app.post("/api/projects/{project_id}/activate")
# async def activate_project(project_id: str):
#     """Set a project as the active project"""
#     project = db.get_project_by_id(project_id)
#     if not project:
#         raise HTTPException(status_code=404, detail="Project not found")
    
#     db.set_active_project(project_id)
#     return {"message": "Project activated successfully"}

# # Project notes endpoints
# @app.get("/api/projects/{project_id}/notes")
# async def get_project_notes(project_id: str):
#     """Get all notes for a project"""
#     project = db.get_project_by_id(project_id)
#     if not project:
#         raise HTTPException(status_code=404, detail="Project not found")
    
#     notes = db.get_all_notes_for_project(project_id)
#     return {"notes": notes}

# @app.get("/api/projects/{project_id}/notes/{note_date}")
# async def get_note_by_date(project_id: str, note_date: str):
#     """Get a note for a specific date"""
#     project = db.get_project_by_id(project_id)
#     if not project:
#         raise HTTPException(status_code=404, detail="Project not found")
    
#     note = db.get_note_by_date(project_id, note_date)
#     if not note:
#         return {"note": None}
    
#     return {"note": note}

# @app.post("/api/projects/{project_id}/notes")
# async def create_note(project_id: str, note_data: NoteCreate):
#     """Create or update a note for a specific date"""
#     project = db.get_project_by_id(project_id)
#     if not project:
#         raise HTTPException(status_code=404, detail="Project not found")
    
#     # Check if note already exists for this date
#     existing_note = db.get_note_by_date(project_id, note_data.note_date)
    
#     if existing_note:
#         # Update existing note
#         updated_note = db.update_note(existing_note['id'], {'content': note_data.content})
#         return {"note": updated_note, "message": "Note updated successfully"}
#     else:
#         # Create new note
#         note_id = str(uuid.uuid4())
#         now = datetime.now().isoformat()
        
#         note_dict = {
#             'id': note_id,
#             'project_id': project_id,
#             'note_date': note_data.note_date,
#             'content': note_data.content,
#             'created_at': now,
#             'updated_at': now
#         }
        
#         db.create_note(note_dict)
#         return {"note": note_dict, "message": "Note created successfully"}

# @app.put("/api/notes/{note_id}")
# async def update_note(note_id: str, updates: NoteUpdate):
#     """Update an existing note"""
#     updated_note = db.update_note(note_id, {'content': updates.content})
#     if not updated_note:
#         raise HTTPException(status_code=404, detail="Note not found")
    
#     return {"note": updated_note, "message": "Note updated successfully"}

# @app.delete("/api/notes/{note_id}")
# async def delete_note(note_id: str):
#     """Delete a note"""
#     deleted = db.delete_note(note_id)
#     if not deleted:
#         raise HTTPException(status_code=404, detail="Note not found")
    
#     return {"message": "Note deleted successfully"}

# # Task endpoints
# @app.get("/")
# async def read_root():
#     """Serve the main application"""
#     return FileResponse("static/index.html")

# @app.get("/api/tasks")
# async def get_tasks():
#     """Get all tasks for the active project"""
#     project_id = get_current_project_id()
#     tasks_list = db.get_all_tasks(project_id)
#     critical_path = calculate_critical_path(project_id)
    
#     root_tasks = []
#     task_children = {}
    
#     for task in tasks_list:
#         parent_id = task.get('parent_id')
#         if parent_id:
#             if parent_id not in task_children:
#                 task_children[parent_id] = []
#             task_children[parent_id].append(task)
#         else:
#             root_tasks.append(task)
    
#     def add_children_recursive(task):
#         task_id = task['id']
#         if task_id in task_children:
#             task['subtasks'] = task_children[task_id]
#             for child in task['subtasks']:
#                 add_children_recursive(child)
#         else:
#             task['subtasks'] = []
#         return task
    
#     hierarchical_tasks = [add_children_recursive(task) for task in root_tasks]
    
#     return {
#         "tasks": tasks_list,
#         "hierarchical_tasks": hierarchical_tasks,
#         "total_tasks": len(tasks_list),
#         "completed_tasks": len([t for t in tasks_list if t['progress'] >= 100]),
#         "critical_path": critical_path,
#         "project_id": project_id
#     }

# @app.get("/api/tasks/{task_id}")
# async def get_task(task_id: str):
#     """Get a specific task"""
#     task = db.get_task_by_id(task_id)
#     if not task:
#         raise HTTPException(status_code=404, detail="Task not found")
    
#     task_logs = db.get_task_logs(task_id)
    
#     return {
#         "task": task,
#         "logs": task_logs[-10:]
#     }

# @app.post("/api/tasks")
# async def create_task(task_data: TaskCreate):
#     """Create a new task"""
#     project_id = get_current_project_id()
#     task_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     valid_dependencies = validate_dependencies(task_id, task_data.dependencies, project_id)
#     default_color = "#666666" if task_data.is_milestone else "#4285f4"
    
#     task_dict = {
#         'id': task_id,
#         'project_id': project_id,
#         'name': task_data.name,
#         'start_date': task_data.start_date,
#         'end_date': task_data.end_date,
#         'progress': task_data.progress,
#         'color': task_data.color if task_data.color != "#4285f4" else default_color,
#         'dependencies': valid_dependencies,
#         'is_milestone': task_data.is_milestone,
#         'parent_id': task_data.parent_id,
#         'description': task_data.description,
#         'assigned_to': task_data.assigned_to,
#         'priority': task_data.priority,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_task(task_dict)
#     log_action("CREATE", task_id, task_data.name, task_dict, project_id)
    
#     return {"task": task_dict, "message": "Task created successfully"}

# @app.put("/api/tasks/{task_id}")
# async def update_task(task_id: str, updates: TaskUpdate):
#     """Update an existing task"""
#     task = db.get_task_by_id(task_id)
#     if not task:
#         raise HTTPException(status_code=404, detail="Task not found")
    
#     old_data = task.copy()
#     update_data = updates.model_dump(exclude_unset=True)
    
#     if 'dependencies' in update_data:
#         update_data['dependencies'] = validate_dependencies(task_id, update_data['dependencies'], task['project_id'])
    
#     updated_task = db.update_task(task_id, update_data)
    
#     log_action("UPDATE", task_id, updated_task['name'], {
#         "old": old_data,
#         "new": updated_task,
#         "changes": update_data
#     }, task['project_id'])
    
#     return {"task": updated_task, "message": "Task updated successfully"}

# @app.delete("/api/tasks/{task_id}")
# async def delete_task(task_id: str):
#     """Delete a task"""
#     task = db.get_task_by_id(task_id)
#     if not task:
#         raise HTTPException(status_code=404, detail="Task not found")
    
#     task_name = task['name']
#     project_id = task['project_id']
    
#     log_action("DELETE", task_id, task_name, task, project_id)
#     db.delete_task(task_id)
    
#     return {"message": "Task deleted successfully"}

# @app.post("/api/tasks/{parent_id}/subtask")
# async def create_subtask(parent_id: str, task_data: TaskCreate):
#     """Create a new subtask under a parent task"""
#     parent_task = db.get_task_by_id(parent_id)
#     if not parent_task:
#         raise HTTPException(status_code=404, detail="Parent task not found")
    
#     if parent_task['is_milestone']:
#         raise HTTPException(status_code=400, detail="Cannot create subtasks under milestones")
    
#     task_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
#     project_id = parent_task['project_id']
    
#     valid_dependencies = validate_dependencies(task_id, task_data.dependencies, project_id)
#     default_color = "#666666" if task_data.is_milestone else "#4285f4"
    
#     task_dict = {
#         'id': task_id,
#         'project_id': project_id,
#         'name': task_data.name,
#         'start_date': task_data.start_date,
#         'end_date': task_data.end_date,
#         'progress': task_data.progress,
#         'color': task_data.color if task_data.color != "#4285f4" else default_color,
#         'dependencies': valid_dependencies,
#         'is_milestone': task_data.is_milestone,
#         'parent_id': parent_id,
#         'description': task_data.description,
#         'assigned_to': task_data.assigned_to,
#         'priority': task_data.priority,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_task(task_dict)
#     log_action("CREATE_SUBTASK", task_id, task_data.name, {
#         **task_dict,
#         "parent_task_name": parent_task['name']
#     }, project_id)
    
#     return {"task": task_dict, "message": f"Subtask created under '{parent_task['name']}'"}

# @app.get("/api/logs")
# async def get_logs(limit: int = 50):
#     """Get recent action logs for active project"""
#     project_id = get_current_project_id()
#     logs = db.get_logs(project_id, limit)
#     return {"logs": logs}

# @app.get("/api/analytics")
# async def get_analytics():
#     """Get project analytics and statistics"""
#     project_id = get_current_project_id()
#     tasks_list = db.get_all_tasks(project_id)
    
#     if not tasks_list:
#         return {"message": "No tasks available for analytics"}
    
#     total_tasks = len(tasks_list)
#     completed_tasks = len([t for t in tasks_list if t['progress'] >= 100])
#     in_progress_tasks = len([t for t in tasks_list if 0 < t['progress'] < 100])
#     not_started_tasks = len([t for t in tasks_list if t['progress'] == 0])
#     milestones = len([t for t in tasks_list if t['is_milestone']])
    
#     avg_progress = sum(t['progress'] for t in tasks_list) / total_tasks
    
#     priority_dist = {}
#     for task in tasks_list:
#         priority_dist[task['priority']] = priority_dist.get(task['priority'], 0) + 1
    
#     critical_path = calculate_critical_path(project_id)
    
#     return {
#         "total_tasks": total_tasks,
#         "completed_tasks": completed_tasks,
#         "in_progress_tasks": in_progress_tasks,
#         "not_started_tasks": not_started_tasks,
#         "milestones": milestones,
#         "completion_rate": (completed_tasks / total_tasks) * 100 if total_tasks > 0 else 0,
#         "average_progress": avg_progress,
#         "priority_distribution": priority_dist,
#         "critical_path_length": len(critical_path),
#         "critical_path": critical_path
#     }

# @app.get("/api/health")
# async def health_check():
#     """Health check endpoint"""
#     project_id = get_current_project_id()
#     tasks_count = len(db.get_all_tasks(project_id))
#     logs_count = len(db.get_logs(project_id, limit=10000))
    
#     return {
#         "status": "healthy",
#         "timestamp": datetime.now().isoformat(),
#         "database": "sqlite",
#         "tasks_count": tasks_count,
#         "logs_count": logs_count,
#         "current_project": project_id
#     }

# # Weekly planner endpoints
# @app.get("/planner")
# async def read_planner():
#     """Serve the weekly planner application"""
#     return FileResponse("static/planner.html")

# @app.post("/api/planners")
# async def create_planner(planner_data: PlannerCreate):
#     """Create a new weekly planner"""
#     planner_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     # Calculate week end date (6 days after start)
#     start_date = datetime.fromisoformat(planner_data.week_start_date)
#     end_date = start_date + timedelta(days=6)
    
#     planner_dict = {
#         'id': planner_id,
#         'project_id': planner_data.project_id or get_current_project_id(),
#         'week_start_date': planner_data.week_start_date,
#         'week_end_date': end_date.date().isoformat(),
#         'custom_rows': [],
#         'custom_columns': [],
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_weekly_planner(planner_dict)
#     return {"planner": planner_dict, "message": "Planner created successfully"}

# @app.get("/api/planners")
# async def get_planners():
#     """Get all weekly planners for the active project"""
#     project_id = get_current_project_id()
#     planners = db.get_all_planners(project_id)
#     return {"planners": planners}

# @app.get("/api/planners/week/{week_start_date}")
# async def get_planner_by_week(week_start_date: str):
#     """Get planner for a specific week"""
#     project_id = get_current_project_id()
#     planner = db.get_planner_by_week(week_start_date, project_id)
    
#     if not planner:
#         # Create a new planner for this week
#         planner_id = str(uuid.uuid4())
#         now = datetime.now().isoformat()
#         start_date = datetime.fromisoformat(week_start_date)
#         end_date = start_date + timedelta(days=6)
        
#         planner = {
#             'id': planner_id,
#             'project_id': project_id,
#             'week_start_date': week_start_date,
#             'week_end_date': end_date.date().isoformat(),
#             'custom_rows': [],
#             'custom_columns': [],
#             'created_at': now,
#             'updated_at': now
#         }
#         db.create_weekly_planner(planner)
    
#     # Get time blocks for this planner
#     time_blocks = db.get_time_blocks(planner['id'])
    
#     return {
#         "planner": planner,
#         "time_blocks": time_blocks
#     }

# @app.put("/api/planners/{planner_id}")
# async def update_planner(planner_id: str, updates: PlannerUpdate):
#     """Update a weekly planner"""
#     update_data = updates.model_dump(exclude_unset=True)
#     updated_planner = db.update_planner(planner_id, update_data)
    
#     if not updated_planner:
#         raise HTTPException(status_code=404, detail="Planner not found")
    
#     return {"planner": updated_planner, "message": "Planner updated successfully"}

# # Time block endpoints
# @app.post("/api/planners/{planner_id}/blocks")
# async def create_time_block(planner_id: str, block_data: TimeBlockCreate):
#     """Create a new time block"""
#     block_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     block_dict = {
#         'id': block_id,
#         'planner_id': planner_id,
#         'day_index': block_data.day_index,
#         'time_slot': block_data.time_slot,
#         'title': block_data.title,
#         'description': block_data.description,
#         'color': block_data.color,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_time_block(block_dict)
#     return {"block": block_dict, "message": "Time block created successfully"}

# @app.get("/api/planners/{planner_id}/blocks")
# async def get_planner_blocks(planner_id: str):
#     """Get all time blocks for a planner"""
#     blocks = db.get_time_blocks(planner_id)
#     return {"blocks": blocks}

# @app.put("/api/blocks/{block_id}")
# async def update_block(block_id: str, updates: TimeBlockUpdate):
#     """Update a time block"""
#     update_data = updates.model_dump(exclude_unset=True)
#     updated_block = db.update_time_block(block_id, update_data)
    
#     if not updated_block:
#         raise HTTPException(status_code=404, detail="Time block not found")
    
#     return {"block": updated_block, "message": "Time block updated successfully"}

# @app.delete("/api/blocks/{block_id}")
# async def delete_block(block_id: str):
#     """Delete a time block"""
#     deleted = db.delete_time_block(block_id)
#     if not deleted:
#         raise HTTPException(status_code=404, detail="Time block not found")
    
#     return {"message": "Time block deleted successfully"}

# # XLSX file endpoints
# @app.post("/api/xlsx/upload")
# async def upload_xlsx(file: UploadFile = File(...)):
#     """Upload an xlsx file"""
#     if not file.filename.endswith(('.xlsx', '.xls')):
#         raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
#     project_id = get_current_project_id()
#     file_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     file_data_bytes = await file.read()
    
#     file_dict = {
#         'id': file_id,
#         'project_id': project_id,
#         'filename': file.filename,
#         'file_data': file_data_bytes,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_xlsx_file(file_dict)
#     return {"file_id": file_id, "filename": file.filename, "message": "File uploaded successfully"}

# @app.get("/api/xlsx")
# async def get_xlsx_files():
#     """Get all xlsx files for the active project"""
#     project_id = get_current_project_id()
#     files = db.get_all_xlsx_files(project_id)
#     return {"files": files}

# @app.get("/api/xlsx/{file_id}/download")
# async def download_xlsx(file_id: str):
#     """Download an xlsx file"""
#     file_data = db.get_xlsx_file(file_id)
#     if not file_data:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return Response(
#         content=file_data['file_data'],
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={
#             "Content-Disposition": f"attachment; filename={file_data['filename']}"
#         }
#     )

# @app.delete("/api/xlsx/{file_id}")
# async def delete_xlsx(file_id: str):
#     """Delete an xlsx file"""
#     deleted = db.delete_xlsx_file(file_id)
#     if not deleted:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return {"message": "File deleted successfully"}

# # Markdown file endpoints
# @app.post("/api/markdown/upload")
# async def upload_markdown(file: UploadFile = File(...)):
#     """Upload a markdown file"""
#     if not file.filename.endswith('.md'):
#         raise HTTPException(status_code=400, detail="Only .md files are allowed")
    
#     project_id = get_current_project_id()
#     file_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     content = (await file.read()).decode('utf-8')
    
#     file_dict = {
#         'id': file_id,
#         'project_id': project_id,
#         'filename': file.filename,
#         'content': content,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_markdown_file(file_dict)
#     return {"file_id": file_id, "filename": file.filename, "message": "File uploaded successfully"}

# @app.get("/api/markdown")
# async def get_markdown_files():
#     """Get all markdown files for the active project"""
#     project_id = get_current_project_id()
#     files = db.get_all_markdown_files(project_id)
#     return {"files": files}

# @app.get("/api/markdown/{file_id}")
# async def get_markdown(file_id: str):
#     """Get a markdown file"""
#     file_data = db.get_markdown_file(file_id)
#     if not file_data:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return {"file": file_data}

# @app.put("/api/markdown/{file_id}")
# async def update_markdown(file_id: str, content: dict):
#     """Update a markdown file"""
#     updated = db.update_markdown_file(file_id, content['content'])
#     if not updated:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return {"message": "File updated successfully", "file": updated}

# @app.delete("/api/markdown/{file_id}")
# async def delete_markdown(file_id: str):
#     """Delete a markdown file"""
#     deleted = db.delete_markdown_file(file_id)
#     if not deleted:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return {"message": "File deleted successfully"}

# # PDF file endpoints
# @app.post("/api/pdf/upload")
# async def upload_pdf(file: UploadFile = File(...)):
#     """Upload a PDF file"""
#     if not file.filename.endswith('.pdf'):
#         raise HTTPException(status_code=400, detail="Only .pdf files are allowed")
    
#     project_id = get_current_project_id()
#     file_id = str(uuid.uuid4())
#     now = datetime.now().isoformat()
    
#     file_data_bytes = await file.read()
    
#     file_dict = {
#         'id': file_id,
#         'project_id': project_id,
#         'filename': file.filename,
#         'file_data': file_data_bytes,
#         'created_at': now,
#         'updated_at': now
#     }
    
#     db.create_pdf_file(file_dict)
#     return {"file_id": file_id, "filename": file.filename, "message": "File uploaded successfully"}

# @app.get("/api/pdf")
# async def get_pdf_files():
#     """Get all PDF files for the active project"""
#     project_id = get_current_project_id()
#     files = db.get_all_pdf_files(project_id)
#     return {"files": files}

# @app.get("/api/pdf/{file_id}/view")
# async def view_pdf(file_id: str):
#     """View a PDF file"""
#     file_data = db.get_pdf_file(file_id)
#     if not file_data:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return Response(
#         content=file_data['file_data'],
#         media_type="application/pdf",
#         headers={
#             "Content-Disposition": f"inline; filename={file_data['filename']}"
#         }
#     )

# @app.delete("/api/pdf/{file_id}")
# async def delete_pdf(file_id: str):
#     """Delete a PDF file"""
#     deleted = db.delete_pdf_file(file_id)
#     if not deleted:
#         raise HTTPException(status_code=404, detail="File not found")
    
#     return {"message": "File deleted successfully"}

# app.mount("/static", StaticFiles(directory="static"), name="static")

# if __name__ == "__main__":
#     import uvicorn
#     print("Starting Gantt Chart Application with SQLite...")
#     print("Access the application at: http://localhost:8000")
#     uvicorn.run(app, host="0.0.0.0", port=8000)

from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime, timedelta
import uuid
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import database as db

app = FastAPI(title="Gantt Chart API", description="Multi-project Gantt chart application with notes")

# Data models
class Project(BaseModel):
    id: str
    name: str
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    color: str = "#4285f4"
    is_active: bool = True
    created_at: str
    updated_at: str

class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    color: str = "#4285f4"

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    color: Optional[str] = None

class Note(BaseModel):
    id: str
    project_id: str
    note_date: str
    content: str
    created_at: str
    updated_at: str

class NoteCreate(BaseModel):
    note_date: str
    content: str

class NoteUpdate(BaseModel):
    content: str

class Task(BaseModel):
    id: str
    project_id: str
    name: str
    start_date: str
    end_date: str
    progress: float = 0.0
    color: str = "#4285f4"
    dependencies: List[str] = []
    is_milestone: bool = False
    parent_id: Optional[str] = None
    created_at: str
    updated_at: str
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    priority: str = "medium"

class TaskCreate(BaseModel):
    name: str
    start_date: str
    end_date: str
    progress: float = 0.0
    color: str = "#4285f4"
    dependencies: List[str] = []
    is_milestone: bool = False
    parent_id: Optional[str] = None
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    priority: str = "medium"

class TaskUpdate(BaseModel):
    name: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    progress: Optional[float] = None
    color: Optional[str] = None
    dependencies: Optional[List[str]] = None
    is_milestone: Optional[bool] = None
    description: Optional[str] = None
    assigned_to: Optional[str] = None
    priority: Optional[str] = None
    parent_id: Optional[str] = None

class WeeklyPlanner(BaseModel):
    id: str
    project_id: Optional[str] = None
    week_start_date: str
    week_end_date: str
    custom_rows: List[str] = []
    custom_columns: List[str] = []
    created_at: str
    updated_at: str

class PlannerCreate(BaseModel):
    week_start_date: str
    project_id: Optional[str] = None

class PlannerUpdate(BaseModel):
    custom_rows: Optional[List[str]] = None
    custom_columns: Optional[List[str]] = None

class TimeBlock(BaseModel):
    id: str
    planner_id: str
    day_index: int
    time_slot: str
    title: Optional[str] = None
    description: Optional[str] = None
    color: str = "#4285f4"
    created_at: str
    updated_at: str

class TimeBlockCreate(BaseModel):
    day_index: int
    time_slot: str
    title: Optional[str] = None
    description: Optional[str] = None
    color: str = "#4285f4"

class TimeBlockUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None

@app.on_event("startup")
async def startup_event():
    """Initialize the database when the app starts"""
    db.init_database()
    print("✓ Database ready")

def get_current_project_id():
    """Get the currently active project ID"""
    project = db.get_active_project()
    if project:
        return project['id']
    
    # If no active project, get the first one or create default
    projects = db.get_all_projects()
    if projects:
        db.set_active_project(projects[0]['id'])
        return projects[0]['id']
    
    # Create default project
    now = datetime.now().isoformat()
    default_project = {
        'id': str(uuid.uuid4()),
        'name': 'My First Project',
        'description': 'Default project',
        'created_at': now,
        'updated_at': now
    }
    db.create_project(default_project)
    db.set_active_project(default_project['id'])
    return default_project['id']

def validate_dependencies(task_id: str, dependencies: List[str], project_id: str) -> List[str]:
    """Validate and filter dependencies to prevent circular references"""
    valid_deps = []
    all_tasks = {task['id']: task for task in db.get_all_tasks(project_id)}
    
    def has_circular_dependency(current_id: str, target_id: str, visited: set = None) -> bool:
        if visited is None:
            visited = set()
        
        if current_id in visited:
            return True
        
        if current_id == target_id:
            return True
        
        visited.add(current_id)
        
        current_task = all_tasks.get(current_id)
        if current_task:
            for dep in current_task['dependencies']:
                if has_circular_dependency(dep, target_id, visited.copy()):
                    return True
        
        return False
    
    for dep_id in dependencies:
        if dep_id in all_tasks and not has_circular_dependency(dep_id, task_id):
            valid_deps.append(dep_id)
    
    return valid_deps

def calculate_critical_path(project_id: str) -> List[str]:
    """Calculate the critical path through the project"""
    all_tasks = {task['id']: task for task in db.get_all_tasks(project_id)}
    
    def get_chain_length(task_id: str, visited: set = None) -> int:
        if visited is None:
            visited = set()
        
        if task_id in visited:
            return 0
        
        visited.add(task_id)
        task = all_tasks.get(task_id)
        if not task:
            return 0
        
        if not task['dependencies']:
            return 1
        
        max_length = 0
        for dep_id in task['dependencies']:
            length = get_chain_length(dep_id, visited.copy())
            max_length = max(max_length, length)
        
        return max_length + 1
    
    max_length = 0
    critical_task = None
    
    for task_id in all_tasks:
        length = get_chain_length(task_id)
        if length > max_length:
            max_length = length
            critical_task = task_id
    
    critical_path = []
    if critical_task:
        def trace_path(task_id: str):
            critical_path.append(task_id)
            task = all_tasks.get(task_id)
            if task and task['dependencies']:
                longest_dep = None
                max_dep_length = 0
                for dep_id in task['dependencies']:
                    dep_length = get_chain_length(dep_id)
                    if dep_length > max_dep_length:
                        max_dep_length = dep_length
                        longest_dep = dep_id
                
                if longest_dep:
                    trace_path(longest_dep)
        
        trace_path(critical_task)
    
    return critical_path

def log_action(action: str, task_id: str, task_name: str, details: Dict, project_id: str, user: str = "system"):
    """Add an action to the log"""
    log_data = {
        'id': str(uuid.uuid4()),
        'project_id': project_id,
        'action': action,
        'task_id': task_id,
        'task_name': task_name,
        'timestamp': datetime.now().isoformat(),
        'details': details,
        'user': user
    }
    db.create_log(log_data)

@app.middleware("http")
async def add_cors_header(request: Request, call_next):
    response = await call_next(request)
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "*"
    return response

# Project endpoints
@app.get("/api/projects")
async def get_projects():
    """Get all projects"""
    projects = db.get_all_projects()
    return {"projects": projects}

@app.get("/api/projects/active")
async def get_active_project():
    """Get the currently active project"""
    project = db.get_active_project()
    if not project:
        raise HTTPException(status_code=404, detail="No active project")
    return {"project": project}

@app.post("/api/projects")
async def create_project(project_data: ProjectCreate):
    """Create a new project"""
    project_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    project_dict = {
        'id': project_id,
        'name': project_data.name,
        'description': project_data.description,
        'start_date': project_data.start_date,
        'end_date': project_data.end_date,
        'color': project_data.color,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_project(project_dict)
    db.set_active_project(project_id)
    
    return {"project": project_dict, "message": "Project created successfully"}

@app.put("/api/projects/{project_id}")
async def update_project(project_id: str, updates: ProjectUpdate):
    """Update an existing project"""
    project = db.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_data = updates.model_dump(exclude_unset=True)
    updated_project = db.update_project(project_id, update_data)
    
    return {"project": updated_project, "message": "Project updated successfully"}

@app.delete("/api/projects/{project_id}")
async def delete_project(project_id: str):
    """Delete a project and all its data"""
    project = db.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if it's the only project
    all_projects = db.get_all_projects()
    if len(all_projects) == 1:
        raise HTTPException(status_code=400, detail="Cannot delete the only project")
    
    db.delete_project(project_id)
    
    # Set another project as active if this was the active one
    if project.get('is_active'):
        remaining_projects = db.get_all_projects()
        if remaining_projects:
            db.set_active_project(remaining_projects[0]['id'])
    
    return {"message": "Project deleted successfully"}

@app.post("/api/projects/{project_id}/activate")
async def activate_project(project_id: str):
    """Set a project as the active project"""
    project = db.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    db.set_active_project(project_id)
    return {"message": "Project activated successfully"}

# Project notes endpoints
@app.get("/api/projects/{project_id}/notes")
async def get_project_notes(project_id: str):
    """Get all notes for a project"""
    project = db.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    notes = db.get_all_notes_for_project(project_id)
    return {"notes": notes}

@app.get("/api/projects/{project_id}/notes/{note_date}")
async def get_note_by_date(project_id: str, note_date: str):
    """Get a note for a specific date"""
    project = db.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    note = db.get_note_by_date(project_id, note_date)
    if not note:
        return {"note": None}
    
    return {"note": note}

@app.post("/api/projects/{project_id}/notes")
async def create_note(project_id: str, note_data: NoteCreate):
    """Create or update a note for a specific date"""
    project = db.get_project_by_id(project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if note already exists for this date
    existing_note = db.get_note_by_date(project_id, note_data.note_date)
    
    if existing_note:
        # Update existing note
        updated_note = db.update_note(existing_note['id'], {'content': note_data.content})
        return {"note": updated_note, "message": "Note updated successfully"}
    else:
        # Create new note
        note_id = str(uuid.uuid4())
        now = datetime.now().isoformat()
        
        note_dict = {
            'id': note_id,
            'project_id': project_id,
            'note_date': note_data.note_date,
            'content': note_data.content,
            'created_at': now,
            'updated_at': now
        }
        
        db.create_note(note_dict)
        return {"note": note_dict, "message": "Note created successfully"}

@app.put("/api/notes/{note_id}")
async def update_note(note_id: str, updates: NoteUpdate):
    """Update an existing note"""
    updated_note = db.update_note(note_id, {'content': updates.content})
    if not updated_note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    return {"note": updated_note, "message": "Note updated successfully"}

@app.delete("/api/notes/{note_id}")
async def delete_note(note_id: str):
    """Delete a note"""
    deleted = db.delete_note(note_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Note not found")
    
    return {"message": "Note deleted successfully"}

# Task endpoints
@app.get("/")
async def read_root():
    """Serve the main application"""
    return FileResponse("static/index.html")

@app.get("/api/tasks")
async def get_tasks():
    """Get all tasks for the active project"""
    project_id = get_current_project_id()
    tasks_list = db.get_all_tasks(project_id)
    critical_path = calculate_critical_path(project_id)
    
    root_tasks = []
    task_children = {}
    
    for task in tasks_list:
        parent_id = task.get('parent_id')
        if parent_id:
            if parent_id not in task_children:
                task_children[parent_id] = []
            task_children[parent_id].append(task)
        else:
            root_tasks.append(task)
    
    def add_children_recursive(task):
        task_id = task['id']
        if task_id in task_children:
            task['subtasks'] = task_children[task_id]
            for child in task['subtasks']:
                add_children_recursive(child)
        else:
            task['subtasks'] = []
        return task
    
    hierarchical_tasks = [add_children_recursive(task) for task in root_tasks]
    
    return {
        "tasks": tasks_list,
        "hierarchical_tasks": hierarchical_tasks,
        "total_tasks": len(tasks_list),
        "completed_tasks": len([t for t in tasks_list if t['progress'] >= 100]),
        "critical_path": critical_path,
        "project_id": project_id
    }

@app.get("/api/tasks/{task_id}")
async def get_task(task_id: str):
    """Get a specific task"""
    task = db.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task_logs = db.get_task_logs(task_id)
    
    return {
        "task": task,
        "logs": task_logs[-10:]
    }

@app.post("/api/tasks")
async def create_task(task_data: TaskCreate):
    """Create a new task"""
    project_id = get_current_project_id()
    task_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    valid_dependencies = validate_dependencies(task_id, task_data.dependencies, project_id)
    default_color = "#666666" if task_data.is_milestone else "#4285f4"
    
    task_dict = {
        'id': task_id,
        'project_id': project_id,
        'name': task_data.name,
        'start_date': task_data.start_date,
        'end_date': task_data.end_date,
        'progress': task_data.progress,
        'color': task_data.color if task_data.color != "#4285f4" else default_color,
        'dependencies': valid_dependencies,
        'is_milestone': task_data.is_milestone,
        'parent_id': task_data.parent_id,
        'description': task_data.description,
        'assigned_to': task_data.assigned_to,
        'priority': task_data.priority,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_task(task_dict)
    log_action("CREATE", task_id, task_data.name, task_dict, project_id)
    
    return {"task": task_dict, "message": "Task created successfully"}

@app.put("/api/tasks/{task_id}")
async def update_task(task_id: str, updates: TaskUpdate):
    """Update an existing task"""
    task = db.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    old_data = task.copy()
    update_data = updates.model_dump(exclude_unset=True)
    
    if 'dependencies' in update_data:
        update_data['dependencies'] = validate_dependencies(task_id, update_data['dependencies'], task['project_id'])
    
    updated_task = db.update_task(task_id, update_data)
    
    log_action("UPDATE", task_id, updated_task['name'], {
        "old": old_data,
        "new": updated_task,
        "changes": update_data
    }, task['project_id'])
    
    return {"task": updated_task, "message": "Task updated successfully"}

@app.delete("/api/tasks/{task_id}")
async def delete_task(task_id: str):
    """Delete a task"""
    task = db.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    
    task_name = task['name']
    project_id = task['project_id']
    
    log_action("DELETE", task_id, task_name, task, project_id)
    db.delete_task(task_id)
    
    return {"message": "Task deleted successfully"}

@app.post("/api/tasks/{parent_id}/subtask")
async def create_subtask(parent_id: str, task_data: TaskCreate):
    """Create a new subtask under a parent task"""
    parent_task = db.get_task_by_id(parent_id)
    if not parent_task:
        raise HTTPException(status_code=404, detail="Parent task not found")
    
    if parent_task['is_milestone']:
        raise HTTPException(status_code=400, detail="Cannot create subtasks under milestones")
    
    task_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    project_id = parent_task['project_id']
    
    valid_dependencies = validate_dependencies(task_id, task_data.dependencies, project_id)
    default_color = "#666666" if task_data.is_milestone else "#4285f4"
    
    task_dict = {
        'id': task_id,
        'project_id': project_id,
        'name': task_data.name,
        'start_date': task_data.start_date,
        'end_date': task_data.end_date,
        'progress': task_data.progress,
        'color': task_data.color if task_data.color != "#4285f4" else default_color,
        'dependencies': valid_dependencies,
        'is_milestone': task_data.is_milestone,
        'parent_id': parent_id,
        'description': task_data.description,
        'assigned_to': task_data.assigned_to,
        'priority': task_data.priority,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_task(task_dict)
    log_action("CREATE_SUBTASK", task_id, task_data.name, {
        **task_dict,
        "parent_task_name": parent_task['name']
    }, project_id)
    
    return {"task": task_dict, "message": f"Subtask created under '{parent_task['name']}'"}

@app.get("/api/logs")
async def get_logs(limit: int = 50):
    """Get recent action logs for active project"""
    project_id = get_current_project_id()
    logs = db.get_logs(project_id, limit)
    return {"logs": logs}

@app.get("/api/analytics")
async def get_analytics():
    """Get project analytics and statistics"""
    project_id = get_current_project_id()
    tasks_list = db.get_all_tasks(project_id)
    
    if not tasks_list:
        return {"message": "No tasks available for analytics"}
    
    total_tasks = len(tasks_list)
    completed_tasks = len([t for t in tasks_list if t['progress'] >= 100])
    in_progress_tasks = len([t for t in tasks_list if 0 < t['progress'] < 100])
    not_started_tasks = len([t for t in tasks_list if t['progress'] == 0])
    milestones = len([t for t in tasks_list if t['is_milestone']])
    
    avg_progress = sum(t['progress'] for t in tasks_list) / total_tasks
    
    priority_dist = {}
    for task in tasks_list:
        priority_dist[task['priority']] = priority_dist.get(task['priority'], 0) + 1
    
    critical_path = calculate_critical_path(project_id)
    
    return {
        "total_tasks": total_tasks,
        "completed_tasks": completed_tasks,
        "in_progress_tasks": in_progress_tasks,
        "not_started_tasks": not_started_tasks,
        "milestones": milestones,
        "completion_rate": (completed_tasks / total_tasks) * 100 if total_tasks > 0 else 0,
        "average_progress": avg_progress,
        "priority_distribution": priority_dist,
        "critical_path_length": len(critical_path),
        "critical_path": critical_path
    }

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    project_id = get_current_project_id()
    tasks_count = len(db.get_all_tasks(project_id))
    logs_count = len(db.get_logs(project_id, limit=10000))
    
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "database": "sqlite",
        "tasks_count": tasks_count,
        "logs_count": logs_count,
        "current_project": project_id
    }

# Weekly planner endpoints
@app.get("/planner")
async def read_planner():
    """Serve the weekly planner application"""
    return FileResponse("static/planner.html")

@app.post("/api/planners")
async def create_planner(planner_data: PlannerCreate):
    """Create a new weekly planner"""
    planner_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    # Calculate week end date (6 days after start)
    start_date = datetime.fromisoformat(planner_data.week_start_date)
    end_date = start_date + timedelta(days=6)
    
    planner_dict = {
        'id': planner_id,
        'project_id': planner_data.project_id or get_current_project_id(),
        'week_start_date': planner_data.week_start_date,
        'week_end_date': end_date.date().isoformat(),
        'custom_rows': [],
        'custom_columns': [],
        'created_at': now,
        'updated_at': now
    }
    
    db.create_weekly_planner(planner_dict)
    return {"planner": planner_dict, "message": "Planner created successfully"}

@app.get("/api/planners")
async def get_planners():
    """Get all weekly planners for the active project"""
    project_id = get_current_project_id()
    planners = db.get_all_planners(project_id)
    return {"planners": planners}

@app.get("/api/planners/week/{week_start_date}")
async def get_planner_by_week(week_start_date: str):
    """Get planner for a specific week (ISO week aligned - Monday start)"""
    project_id = get_current_project_id()
    
    # Parse the provided date and align to ISO week (Monday start)
    try:
        start_date = datetime.fromisoformat(week_start_date).date()
        
        # Calculate ISO week Monday (day of week: Monday=0, Sunday=6)
        days_since_monday = start_date.weekday()
        iso_monday = start_date - timedelta(days=days_since_monday)
        
        # Use ISO Monday as the week start
        week_start_date = iso_monday.isoformat()
        
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    planner = db.get_planner_by_week(week_start_date, project_id)
    
    if not planner:
        # Create a new planner for this week
        planner_id = str(uuid.uuid4())
        now = datetime.now().isoformat()
        start_date = datetime.fromisoformat(week_start_date).date()
        end_date = start_date + timedelta(days=6)
        
        planner = {
            'id': planner_id,
            'project_id': project_id,
            'week_start_date': week_start_date,
            'week_end_date': end_date.isoformat(),
            'custom_rows': [],
            'custom_columns': [],
            'created_at': now,
            'updated_at': now
        }
        db.create_weekly_planner(planner)
    
    # Get time blocks for this planner
    time_blocks = db.get_time_blocks(planner['id'])
    
    return {
        "planner": planner,
        "time_blocks": time_blocks
    }

@app.put("/api/planners/{planner_id}")
async def update_planner(planner_id: str, updates: PlannerUpdate):
    """Update a weekly planner"""
    update_data = updates.model_dump(exclude_unset=True)
    updated_planner = db.update_planner(planner_id, update_data)
    
    if not updated_planner:
        raise HTTPException(status_code=404, detail="Planner not found")
    
    return {"planner": updated_planner, "message": "Planner updated successfully"}

# Time block endpoints
@app.post("/api/planners/{planner_id}/blocks")
async def create_time_block(planner_id: str, block_data: TimeBlockCreate):
    """Create a new time block"""
    block_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    block_dict = {
        'id': block_id,
        'planner_id': planner_id,
        'day_index': block_data.day_index,
        'time_slot': block_data.time_slot,
        'title': block_data.title,
        'description': block_data.description,
        'color': block_data.color,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_time_block(block_dict)
    return {"block": block_dict, "message": "Time block created successfully"}

@app.get("/api/planners/{planner_id}/blocks")
async def get_planner_blocks(planner_id: str):
    """Get all time blocks for a planner"""
    blocks = db.get_time_blocks(planner_id)
    return {"blocks": blocks}

@app.put("/api/blocks/{block_id}")
async def update_block(block_id: str, updates: TimeBlockUpdate):
    """Update a time block"""
    update_data = updates.model_dump(exclude_unset=True)
    updated_block = db.update_time_block(block_id, update_data)
    
    if not updated_block:
        raise HTTPException(status_code=404, detail="Time block not found")
    
    return {"block": updated_block, "message": "Time block updated successfully"}

@app.delete("/api/blocks/{block_id}")
async def delete_block(block_id: str):
    """Delete a time block"""
    deleted = db.delete_time_block(block_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Time block not found")
    
    return {"message": "Time block deleted successfully"}

# XLSX file endpoints
@app.post("/api/xlsx/upload")
async def upload_xlsx(file: UploadFile = File(...)):
    """Upload an xlsx file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    project_id = get_current_project_id()
    file_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    file_data_bytes = await file.read()
    
    file_dict = {
        'id': file_id,
        'project_id': project_id,
        'filename': file.filename,
        'file_data': file_data_bytes,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_xlsx_file(file_dict)
    return {"file_id": file_id, "filename": file.filename, "message": "File uploaded successfully"}

@app.get("/api/xlsx")
async def get_xlsx_files():
    """Get all xlsx files for the active project"""
    project_id = get_current_project_id()
    files = db.get_all_xlsx_files(project_id)
    return {"files": files}

@app.get("/api/xlsx/{file_id}/download")
async def download_xlsx(file_id: str):
    """Download an xlsx file"""
    file_data = db.get_xlsx_file(file_id)
    if not file_data:
        raise HTTPException(status_code=404, detail="File not found")
    
    return Response(
        content=file_data['file_data'],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={file_data['filename']}"
        }
    )

@app.delete("/api/xlsx/{file_id}")
async def delete_xlsx(file_id: str):
    """Delete an xlsx file"""
    deleted = db.delete_xlsx_file(file_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="File not found")
    
    return {"message": "File deleted successfully"}

# Markdown file endpoints
@app.post("/api/markdown/upload")
async def upload_markdown(file: UploadFile = File(...)):
    """Upload a markdown file"""
    if not file.filename.endswith('.md'):
        raise HTTPException(status_code=400, detail="Only .md files are allowed")
    
    project_id = get_current_project_id()
    file_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    content = (await file.read()).decode('utf-8')
    
    file_dict = {
        'id': file_id,
        'project_id': project_id,
        'filename': file.filename,
        'content': content,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_markdown_file(file_dict)
    return {"file_id": file_id, "filename": file.filename, "message": "File uploaded successfully"}

@app.get("/api/markdown")
async def get_markdown_files():
    """Get all markdown files for the active project"""
    project_id = get_current_project_id()
    files = db.get_all_markdown_files(project_id)
    return {"files": files}

@app.get("/api/markdown/{file_id}")
async def get_markdown(file_id: str):
    """Get a markdown file"""
    file_data = db.get_markdown_file(file_id)
    if not file_data:
        raise HTTPException(status_code=404, detail="File not found")
    
    return {"file": file_data}

@app.put("/api/markdown/{file_id}")
async def update_markdown(file_id: str, content: dict):
    """Update a markdown file"""
    updated = db.update_markdown_file(file_id, content['content'])
    if not updated:
        raise HTTPException(status_code=404, detail="File not found")
    
    return {"message": "File updated successfully", "file": updated}

@app.delete("/api/markdown/{file_id}")
async def delete_markdown(file_id: str):
    """Delete a markdown file"""
    deleted = db.delete_markdown_file(file_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="File not found")
    
    return {"message": "File deleted successfully"}

# PDF file endpoints
@app.post("/api/pdf/upload")
async def upload_pdf(file: UploadFile = File(...)):
    """Upload a PDF file"""
    if not file.filename.endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Only .pdf files are allowed")
    
    project_id = get_current_project_id()
    file_id = str(uuid.uuid4())
    now = datetime.now().isoformat()
    
    file_data_bytes = await file.read()
    
    file_dict = {
        'id': file_id,
        'project_id': project_id,
        'filename': file.filename,
        'file_data': file_data_bytes,
        'created_at': now,
        'updated_at': now
    }
    
    db.create_pdf_file(file_dict)
    return {"file_id": file_id, "filename": file.filename, "message": "File uploaded successfully"}

@app.get("/api/pdf")
async def get_pdf_files():
    """Get all PDF files for the active project"""
    project_id = get_current_project_id()
    files = db.get_all_pdf_files(project_id)
    return {"files": files}

@app.get("/api/pdf/{file_id}/view")
async def view_pdf(file_id: str):
    """View a PDF file"""
    file_data = db.get_pdf_file(file_id)
    if not file_data:
        raise HTTPException(status_code=404, detail="File not found")
    
    return Response(
        content=file_data['file_data'],
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"inline; filename={file_data['filename']}"
        }
    )

@app.delete("/api/pdf/{file_id}")
async def delete_pdf(file_id: str):
    """Delete a PDF file"""
    deleted = db.delete_pdf_file(file_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="File not found")
    
    return {"message": "File deleted successfully"}

# Excel planner export/import endpoints
@app.get("/api/planners/{planner_id}/export")
async def export_planner_to_excel(planner_id: str):
    """Export a weekly planner to Excel format"""
    planner = db.get_planner_by_id(planner_id)
    if not planner:
        raise HTTPException(status_code=404, detail="Planner not found")
    
    time_blocks = db.get_time_blocks(planner_id)
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Planner"
    
    # Define days and time slots
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    time_slots = []
    for hour in range(4, 21):
        for minute in [0, 30]:
            time_slots.append(f"{hour:02d}:{minute:02d}")
    
    # Style definitions
    header_fill = PatternFill(start_color="4285F4", end_color="4285F4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    time_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    ws['A1'] = 'Time'
    ws['A1'].fill = header_fill
    ws['A1'].font = header_font
    ws['A1'].border = border
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx, day in enumerate(days, start=2):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = day
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write time slots and data
    blocks_dict = {}
    for block in time_blocks:
        key = f"{block['day_index']}-{block['time_slot']}"
        blocks_dict[key] = block
    
    for row_idx, time_slot in enumerate(time_slots, start=2):
        # Time column
        time_cell = ws.cell(row=row_idx, column=1)
        time_cell.value = time_slot
        time_cell.fill = time_fill
        time_cell.border = border
        time_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Day columns
        for day_idx in range(7):
            cell = ws.cell(row=row_idx, column=day_idx + 2)
            cell.border = border
            
            key = f"{day_idx}-{time_slot}"
            if key in blocks_dict:
                block = blocks_dict[key]
                cell.value = block['title'] or ''
                if block['description']:
                    cell.comment = openpyxl.comments.Comment(block['description'], "System")
                
                # Apply color
                try:
                    color_hex = block['color'].lstrip('#')
                    cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                except:
                    pass
            
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 12
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Set row heights
    ws.row_dimensions[1].height = 25
    for row in range(2, len(time_slots) + 2):
        ws.row_dimensions[row].height = 40
    
    # Save to bytes
    excel_bytes = io.BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    
    # Generate filename
    start_date = datetime.fromisoformat(planner['week_start_date'])
    filename = f"planner_{start_date.strftime('%Y-W%V')}.xlsx"
    
    return Response(
        content=excel_bytes.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@app.post("/api/planners/{planner_id}/import")
async def import_planner_from_excel(planner_id: str, file: UploadFile = File(...)):
    """Import time blocks from an Excel file into a planner"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files are allowed")
    
    planner = db.get_planner_by_id(planner_id)
    if not planner:
        raise HTTPException(status_code=404, detail="Planner not found")
    
    # Read Excel file
    file_data = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(file_data))
    ws = wb.active
    
    # Parse days from header row
    days_map = {}
    for col_idx in range(2, 9):  # Columns B through H
        day_name = ws.cell(row=1, column=col_idx).value
        if day_name:
            days_map[col_idx] = col_idx - 2  # Map column to day index (0-6)
    
    # Clear existing blocks for this planner
    existing_blocks = db.get_time_blocks(planner_id)
    for block in existing_blocks:
        db.delete_time_block(block['id'])
    
    # Parse time blocks
    imported_count = 0
    for row_idx in range(2, ws.max_row + 1):
        time_slot = ws.cell(row=row_idx, column=1).value
        if not time_slot or not isinstance(time_slot, str):
            continue
        
        # Extract HH:MM format
        time_slot = time_slot.strip()
        if ':' in time_slot:
            time_slot = time_slot.split(':')[0] + ':' + time_slot.split(':')[1][:2]
        
        for col_idx, day_idx in days_map.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                block_id = str(uuid.uuid4())
                now = datetime.now().isoformat()
                
                # Extract color
                color = "#4285f4"  # Default
                if cell.fill and cell.fill.start_color:
                    try:
                        color_hex = str(cell.fill.start_color.rgb)
                        if len(color_hex) == 8:  # ARGB format
                            color = f"#{color_hex[2:]}"
                        elif len(color_hex) == 6:  # RGB format
                            color = f"#{color_hex}"
                    except:
                        pass
                
                # Extract description from comment
                description = None
                if cell.comment:
                    description = cell.comment.text
                
                block_dict = {
                    'id': block_id,
                    'planner_id': planner_id,
                    'day_index': day_idx,
                    'time_slot': time_slot,
                    'title': str(cell.value),
                    'description': description,
                    'color': color,
                    'created_at': now,
                    'updated_at': now
                }
                
                db.create_time_block(block_dict)
                imported_count += 1
    
    return {"message": f"Successfully imported {imported_count} time blocks", "count": imported_count}

@app.get("/api/xlsx/{file_id}/read")
async def read_xlsx_data(file_id: str):
    """Read Excel file data and return as JSON"""
    file_data = db.get_xlsx_file(file_id)
    if not file_data:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Load Excel file
    wb = openpyxl.load_workbook(io.BytesIO(file_data['file_data']))
    
    # Process all sheets
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Read data
        data = []
        for row in ws.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                cell_info = {
                    'value': cell.value,
                    'style': {}
                }
                
                # Extract basic styling
                if cell.fill and cell.fill.start_color:
                    try:
                        color_hex = str(cell.fill.start_color.rgb)
                        if len(color_hex) >= 6:
                            cell_info['style']['background'] = f"#{color_hex[-6:]}"
                    except:
                        pass
                
                if cell.font:
                    if cell.font.bold:
                        cell_info['style']['bold'] = True
                    if cell.font.color:
                        try:
                            color_hex = str(cell.font.color.rgb)
                            if len(color_hex) >= 6:
                                cell_info['style']['color'] = f"#{color_hex[-6:]}"
                        except:
                            pass
                
                row_data.append(cell_info)
            data.append(row_data)
        
        sheets_data[sheet_name] = {
            'data': data,
            'dimensions': {
                'rows': ws.max_row,
                'cols': ws.max_column
            }
        }
    
    return {
        "filename": file_data['filename'],
        "sheets": sheets_data
    }

@app.put("/api/xlsx/{file_id}/update")
async def update_xlsx_data(file_id: str, update_data: Dict[str, Any]):
    """Update Excel file with new data"""
    file_data = db.get_xlsx_file(file_id)
    if not file_data:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Load Excel file
    wb = openpyxl.load_workbook(io.BytesIO(file_data['file_data']))
    
    # Update data
    for sheet_name, sheet_updates in update_data.get('sheets', {}).items():
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        
        for update in sheet_updates.get('cells', []):
            row = update.get('row')
            col = update.get('col')
            value = update.get('value')
            
            if row is not None and col is not None:
                cell = ws.cell(row=row + 1, column=col + 1)  # Excel is 1-indexed
                cell.value = value
                
                # Apply styling if provided
                style = update.get('style', {})
                if 'background' in style:
                    try:
                        color_hex = style['background'].lstrip('#')
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    except:
                        pass
                
                if style.get('bold'):
                    cell.font = Font(bold=True)
    
    # Save updated file
    updated_bytes = io.BytesIO()
    wb.save(updated_bytes)
    updated_bytes.seek(0)
    
    # Update in database
    db.update_xlsx_file(file_id, updated_bytes.getvalue())
    
    return {"message": "File updated successfully"}

# Serve Excel editor page
@app.get("/excel_editor.html")
async def excel_editor():
    """Serve the Excel editor page"""
    return FileResponse("excel_editor.html")

app.mount("/static", StaticFiles(directory="static"), name="static")

if __name__ == "__main__":
    import uvicorn
    print("Starting Gantt Chart Application with SQLite...")
    print("Access the application at: http://localhost:8000")
    uvicorn.run(app, host="0.0.0.0", port=8000)